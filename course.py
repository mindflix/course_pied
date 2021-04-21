import time, json
from openpyxl import Workbook

wb = Workbook()
ws1 = wb.worksheets[0]
output = {"data": []}


class User:
    def __init__(self, user):
        self.user = user
        self.is_json_empty()

    def is_json_empty(self):
        with open("./data/timetable.json", "r+") as file:
            try:
                json_string = json.load(file)
            except (json.JSONDecodeError):
                json.dump(output, file)

    def add_time_slot(self, start, end):
        time_dict = {}
        time_dict["id"] = self.get_last_id() + 1
        time_dict["user"] = self.user
        time_dict["duration"] = [start, end]
        output["data"].append(time_dict)

    def get_last_id(self):
        with open("./data/timetable.json", "r+") as file:
            json_string = json.load(file)
            json_data = json_string["data"]
            try:
                last_id_json = json_data[-1]["id"]
            except (IndexError, KeyError):
                last_id_json = 0

            try:
                last_id_output = output["data"][-1]["id"]
            except (IndexError, KeyError):
                last_id_output = 0

            if last_id_json == 0 and last_id_output == 0:
                return 0
            elif last_id_json > 0 and last_id_output == 0:
                return last_id_json
            elif last_id_json == 0 and last_id_output > 0:
                return last_id_output
            else:
                return last_id_output

    def write_json(self):
        with open("./data/timetable.json", "r+") as file:
            json_string = json.load(file)
            if type(json_string) == dict:
                for i in output["data"]:
                    json_string["data"].append(i)
                file.seek(0)
                json.dump(json_string, file, indent=4)
            else:
                json.dump(output, file, indent=4)


def write_in_cell(cell, value):
    ws1.cell(cell.row, cell.column, value)


def write_in_line(data):
    print(f"Workbook->Writing line for {data}")
    table = ws1.iter_cols(min_row=1, max_col=len(data))
    for value in data:
        write_in_cell(next(table)[0], value)
    print(f"Workbook->Line written for {data}")


def main():
    h1 = User("Nicolas Demol")
    h1.add_time_slot(10, 11)
    h1.write_json()


if __name__ == "__main__":
    main()